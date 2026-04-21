"""
excel_export.py — экспорт данных SOV в Excel.
Требует: openpyxl
"""
import io
import logging
from datetime import datetime

logger = logging.getLogger(__name__)


def export_users_xlsx(users: list) -> bytes:
    """Возвращает байты .xlsx файла со списком участников."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl не установлен. Добавь в requirements.txt")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Участники SOV"

    # Стили
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor="1F4E79")
    center       = Alignment(horizontal="center", vertical="center")
    thin_border  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )
    ban_fill     = PatternFill("solid", fgColor="FFD7D7")
    good_fill    = PatternFill("solid", fgColor="D7FFD7")

    headers = ["№", "ФИО", "Группа", "Пол", "Рейтинг", "Ивентов",
               "Страйк", "Поинты", "Статус", "Язык", "Рефералов", "Дата регистрации"]
    ws.append(headers)

    # Форматируем заголовок
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = thin_border

    # Данные
    for idx, u in enumerate(users, 1):
        ban_status = {"none": "✅ Активен", "temp": "⏳ Временный бан", "full": "🚫 Постоянный бан"}.get(
            u.get("ban_type","none"), "✅ Активен"
        )
        reg_date = ""
        try:
            reg_date = datetime.fromisoformat(u["registered_at"]).strftime("%d.%m.%Y")
        except Exception:
            reg_date = str(u.get("registered_at",""))

        row = [
            idx,
            u.get("full_name",""),
            u.get("group_name",""),
            u.get("gender",""),
            u.get("rating", 0),
            u.get("experience", 0),
            u.get("streak", 0),
            u.get("points", 0),
            ban_status,
            u.get("lang","ru").upper(),
            u.get("referral_count", 0),
            reg_date,
        ]
        ws.append(row)
        row_num = idx + 1
        # Цвет строки по статусу бана
        fill = ban_fill if u.get("ban_type","none") != "none" else None
        for col_num in range(1, len(headers)+1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border    = thin_border
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill

    # Ширина колонок
    col_widths = [4, 30, 12, 6, 9, 9, 8, 8, 18, 7, 10, 16]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.row_dimensions[1].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_events_xlsx(events: list, applications_map: dict = None) -> bytes:
    """
    events: list of event dicts
    applications_map: {event_id: [application dicts]} — если передан, добавляет листы по ивентам
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl не установлен.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ивенты SOV"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    center      = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )
    closed_fill = PatternFill("solid", fgColor="F0F0F0")

    headers = ["ID", "Название", "Дата", "Время", "Место", "Длительность",
               "Всего мест", "♂", "♀", "Строгий пол", "Статус", "Создан"]
    ws.append(headers)
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = center; cell.border = thin_border

    for ev in events:
        created = ""
        try:
            created = datetime.fromisoformat(ev["created_at"]).strftime("%d.%m.%Y")
        except Exception:
            created = str(ev.get("created_at",""))
        row = [
            ev.get("id"),
            ev.get("title",""),
            ev.get("event_date",""),
            ev.get("event_time",""),
            ev.get("location",""),
            ev.get("duration",""),
            ev.get("total_slots",0),
            ev.get("male_slots",0),
            ev.get("female_slots",0),
            "Да" if ev.get("gender_strict") else "Нет",
            "🟢 Открыт" if ev.get("is_active") else "🔴 Закрыт",
            created,
        ]
        ws.append(row)
        row_num = events.index(ev) + 2
        for col_num in range(1, len(headers)+1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if not ev.get("is_active"):
                cell.fill = closed_fill

    col_widths = [5, 30, 12, 8, 25, 14, 10, 5, 5, 12, 12, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Доп. листы по каждому ивенту если переданы заявки
    if applications_map:
        for event_id, apps in applications_map.items():
            if not apps:
                continue
            ev = next((e for e in events if e["id"] == event_id), None)
            if not ev:
                continue
            sheet_title = f"Ивент {event_id}"[:31]
            ws2 = wb.create_sheet(title=sheet_title)
            ws2.append([f"Ивент: {ev['title']}", f"Дата: {ev['event_date']}"])
            ws2.append([])
            sub_headers = ["ФИО", "Группа", "Пол", "Рейтинг", "Опыт", "Статус", "Присутствовал"]
            ws2.append(sub_headers)
            for col_num, _ in enumerate(sub_headers, 1):
                cell = ws2.cell(row=3, column=col_num)
                cell.font = Font(bold=True)
            for app in apps:
                ws2.append([
                    app.get("full_name",""),
                    app.get("group_name",""),
                    app.get("gender",""),
                    app.get("rating",0),
                    app.get("experience",0),
                    {"selected":"Выбран","pending":"Ожидает","rejected":"Отклонён"}.get(app.get("status",""),""),
                    "✅" if app.get("attended") else "—",
                ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
